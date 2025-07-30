import os
import re
import uuid
import logging
from datetime import datetime
from pathlib import Path
from threading import Thread, Event
from typing import Any, Dict, List, Union
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
from . import utils, config, gemini_client
import os
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule


# Configure module-level logger
logger = logging.getLogger(__name__)
if not logger.handlers:
    handler = logging.StreamHandler()
    handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    logger.addHandler(handler)
logger.setLevel(logging.INFO)

# Global dictionary to track active background tasks
# The key is the report_id, the value contains a threading.Event object
ACTIVE_TASKS = {}

def _background_worker(
    companies_df: pd.DataFrame,
    report_id: str,
    report_name: str,
    gemini_api_key: str,
    pe_firms_list: List[str],
    original_filepath: str,
    cancel_event: Event
) -> None:
    """The main background task for running the analysis."""
    try:
        start = datetime.now()
        logger.info('Background worker started for report ID: %s', report_id)
        results = []
        unique_pe = set()
        newly_discovered_pe_firms = set()
        
        company_names = [str(name).strip() for name in companies_df['Company Name'].dropna() if name]

        with ThreadPoolExecutor(max_workers=5) as executor:
            future_to_company = {
                executor.submit(gemini_client.analyze_company, name, gemini_api_key, pe_firms_list, newly_discovered_pe_firms): name
                for name in company_names
            }
            
            for future in as_completed(future_to_company):
                if cancel_event.is_set():
                    logger.info(f"Cancellation signal received for report ID: {report_id}. Stopping analysis.")
                    for f in future_to_company:
                        f.cancel()
                    return

                company_name = future_to_company[future]
                try:
                    data = future.result()
                    results.append(data)
                    if data.get('is_pe_owned') and data.get('pe_owner_names'):
                        unique_pe.update(data.get('pe_owner_names', []))
                except Exception as exc:
                    logger.error('%r generated an exception: %s', company_name, exc)
                    results.append({'company_name': company_name, 'error': f'An exception occurred: {exc}'})

        if cancel_event.is_set():
            logger.info(f"Cancellation signal received for report ID: {report_id} before PE research.")
            return

        if newly_discovered_pe_firms:
            logger.info(f"Discovered {len(newly_discovered_pe_firms)} new PE firms. Updating master list.")
            updated_pe_firms = sorted(list(set(pe_firms_list) | newly_discovered_pe_firms))
            utils.save_pe_firms(updated_pe_firms)
        
        pe_firms_insights = {}
        if unique_pe:
            with ThreadPoolExecutor(max_workers=5) as executor:
                future_to_pe = {executor.submit(gemini_client.research_pe_portfolio, pe_name, gemini_api_key): pe_name for pe_name in unique_pe}
                for future in as_completed(future_to_pe):
                    if cancel_event.is_set():
                        logger.info(f"Cancellation signal received for report ID: {report_id}. Stopping PE research.")
                        for f in future_to_pe:
                            f.cancel()
                        return
                    
                    pe_name = future_to_pe[future]
                    try:
                        pe_firms_insights[pe_name] = future.result()
                    except Exception as exc:
                        logger.error('%r generated an exception during PE research: %s', pe_name, exc)
                        pe_firms_insights[pe_name] = {'name': pe_name, 'error': f'An exception occurred: {exc}'}

        if cancel_event.is_set():
            logger.info(f"Cancellation signal received for report ID: {report_id} before saving.")
            return

        verified_results = _cross_reference_results(results, pe_firms_insights)
        end = datetime.now()
        duration_seconds = (end - start).total_seconds()
        report = {
            'report_name': report_name,
            'data': verified_results,
            'pe_firms_insights': pe_firms_insights,
            'analysis_duration_seconds': duration_seconds,
            'analysis_start_time': start.isoformat(),
            'analysis_end_time': end.isoformat()
        }
        path = Path(config.REPORTS_FOLDER) / f"{report_id}.json"
        utils.save_json_file(str(path), report)
        
        history = utils.load_history()
        for entry in history:
            if entry['id'] == report_id:
                entry.update({
                    'status': 'Completed',
                    'file_path': str(path),
                    'completed_at': datetime.now().isoformat(),
                    'analysis_duration_seconds': duration_seconds
                })
                break
        utils.save_history(history)
        logger.info('Background worker completed for report ID: %s', report_id)
    finally:
        if report_id in ACTIVE_TASKS:
            del ACTIVE_TASKS[report_id]
            logger.info(f"Removed task {report_id} from active tasks list.")

def start_company_analysis(companies_df: pd.DataFrame, original_filepath: str) -> Dict[str, str]:
    gemini_api_key = os.environ.get('GEMINI_API_KEY')
    if not gemini_api_key:
        logger.error('GEMINI_API_KEY is not configured. Please set it in your .env file.')
        return {'error': 'GEMINI_API_KEY is not configured. Please set it in your .env file.'}

    report_id = str(uuid.uuid4())
    report_name = f"Analysis Report - {datetime.now():%Y-%m-%d %H:%M:%S}"

    history = utils.load_history()
    history.insert(0, {
        'id': report_id,
        'name': report_name,
        'date': datetime.now().isoformat(),
        'status': 'Pending',
        'num_companies': len(companies_df),
        'file_path': None,
        'original_filepath': original_filepath,
        'completed_at': None,
        'analysis_duration_seconds': None
    })
    utils.save_history(history)

    cancel_event = Event()
    ACTIVE_TASKS[report_id] = {"cancel_event": cancel_event}

    Thread(
        target=_background_worker,
        args=(companies_df, report_id, report_name, gemini_api_key, utils.load_pe_firms(), original_filepath, cancel_event),
        daemon=True
    ).start()

    logger.info('Analysis started for report ID: %s', report_id)
    return {'message': 'File uploaded and analysis started! You can check the history for status.', 'report_id': report_id, 'report_name': report_name}

def delete_report(report_id: str) -> bool:
    """
    Deletes a report. If the report is pending, it cancels the task first.
    """
    if report_id in ACTIVE_TASKS:
        logger.info(f"Report {report_id} is pending. Sending cancellation signal.")
        ACTIVE_TASKS[report_id]['cancel_event'].set()
        time.sleep(1) # Give the thread a moment to recognize the signal

    history = utils.load_history()
    report_to_delete = next((report for report in history if report['id'] == report_id), None)
    
    if not report_to_delete:
        logger.warning(f"Attempted to delete non-existent report with ID: {report_id}")
        return False

    updated_history = [report for report in history if report['id'] != report_id]

    try:
        if report_to_delete.get('file_path') and os.path.exists(report_to_delete['file_path']):
            os.remove(report_to_delete['file_path'])
        if report_to_delete.get('original_filepath') and os.path.exists(report_to_delete['original_filepath']):
            os.remove(report_to_delete['original_filepath'])
        
        download_filepath = os.path.join(config.REPORTS_FOLDER, f"{report_id}_analysis_results.xlsx")
        if os.path.exists(download_filepath):
            os.remove(download_filepath)

        utils.save_history(updated_history)
        logger.info(f"Successfully deleted report and files for ID: {report_id}")
        return True
    except OSError as e:
        logger.error(f"Error deleting files for report {report_id}: {e}", exc_info=True)
        utils.save_history(updated_history)
        return False
    
def _cross_reference_results(
    company_results: List[Dict[str, Any]],
    pe_insights: Dict[str, Any]
) -> List[Dict[str, Any]]:
        
    logger.info("Starting cross-referencing and verification step...")
    
    # Create a simple lookup map of portfolio companies to their PE owners
    portfolio_to_owner_map = {}
    for pe_name, insights in pe_insights.items():
        if insights.get('portfolio_companies'):
            for company in insights['portfolio_companies']:
                original_name = company.get('name')
                if original_name:
                    normalized_name = _normalize_company_name(original_name)
                    if normalized_name not in portfolio_to_owner_map:
                        portfolio_to_owner_map[normalized_name] = []
                    portfolio_to_owner_map[normalized_name].append(pe_name)

    # Iterate through the original results to find and correct misses
    for company in company_results:
        # We now check any company that isn't already categorized as PE-backed or owned.
        if company.get('ownership_category') not in ['PE-Owned', 'Public (PE-Backed)']:
            original_name = company.get('company_name')
            normalized_name = _normalize_company_name(original_name)
            
            if normalized_name in portfolio_to_owner_map:
                owners = portfolio_to_owner_map[normalized_name]
                logger.warning(
                    f"CORRECTION: Found missed PE relationship for '{original_name}'. "
                    f"Owned by: {', '.join(owners)}."
                )
                # This entry was auto-corrected, so it should be flagged for review.
                company['needs_review'] = True
                review_reason = f"Auto-corrected: Initial analysis missed PE ownership by {', '.join(owners)}."
                
                # Append to an existing review reason if there is one
                if company.get('review_reason'):
                    company['review_reason'] += f" | {review_reason}"
                else:
                    company['review_reason'] = review_reason
                    
                # 1. Update the main flag
                company['is_pe_owned'] = True
                company['flagged_as_pe_account'] = True
                
                # 2. Intelligently set the new category
                if company.get('public_private') == 'Public':
                    company['ownership_category'] = 'Public (PE-Backed)'
                else:
                    company['ownership_category'] = 'PE-Owned'

                # 3. Add the discovered owners, avoiding duplicates
                existing_owners = set(company.get('pe_owner_names', []))
                for owner in owners:
                    existing_owners.add(owner)
                company['pe_owner_names'] = sorted(list(existing_owners))

                # 4. Add a note about the correction
                correction_note = "Ownership corrected by cross-referencing PE portfolio data."
                summary = company.get('ownership_structure', 'N/A')
                if summary and summary != 'N/A':
                    company['ownership_structure'] = f"{summary} NOTE: {correction_note}"
                else:
                    company['ownership_structure'] = correction_note
                    
    logger.info("Cross-referencing finished.")
    return company_results

def _normalize_company_name(name: str) -> str:

    if not isinstance(name, str):
        return ""
        
    name = name.lower()
    
    # List of common suffixes to remove (as whole words)
    suffixes = [
        'inc', 'llc', 'lp', 'ltd', 'gmbh', 'sa', 'ag', 'nv', 'bv',
        'corporation', 'corp', 'company', 'co', 'limited', 'holding', 'holdings'
    ]
    
    # Remove punctuation
    name = re.sub(r'[.,;()]', '', name)
    
    # Remove suffixes
    words = name.split()
    words = [word for word in words if word not in suffixes]
    
    return ' '.join(words).strip()

def create_downloadable_report(report_id: str) -> Union[str, None]:
    history = utils.load_history()
    report_entry = next((item for item in history if item["id"] == report_id), None)

    if not report_entry or not report_entry.get('original_filepath'):
        logger.error(f"Could not find original file path for report ID: {report_id}")
        return None

    original_filepath = report_entry['original_filepath']
    report_json_path = report_entry['file_path']

    try:
        original_df = pd.read_excel(original_filepath)
        report_data = utils.load_json_file(report_json_path)
        analysis_results = report_data.get('data', [])
        results_df = pd.DataFrame(analysis_results)

        merged_df = pd.merge(original_df, results_df, left_on='Company Name', right_on='company_name', how='left')
        
        final_df = merged_df[['Company Name']].copy()
        
        # Add the 'Needs Review' and 'Review Reason' columns
        final_df['Needs Review'] = merged_df['needs_review'].apply(lambda x: 'Yes' if x else 'No')
        final_df['Review Reason'] = merged_df['review_reason']
        final_df['Summary'] = merged_df['ownership_structure']
        final_df['Category'] = merged_df['ownership_category']
        final_df['Status'] = merged_df['public_private']
        final_df['Nation'] = merged_df['nation']
        
        pe_owners = merged_df['pe_owner_names'].dropna()
        owner_cols = []
        if not pe_owners.empty:
            max_owners = pe_owners.apply(len).max()
            if max_owners > 0:
                owner_cols = [f'PE Owner {i+1}' for i in range(max_owners)]
                owners_df = pd.DataFrame(pe_owners.tolist(), index=pe_owners.index).reindex(columns=range(max_owners))
                owners_df.columns = owner_cols
                final_df = final_df.join(owners_df)

        # Sort rows
        final_df.sort_values(by=['Category', 'Company Name'], inplace=True)

        download_filename = f"{report_id}_analysis_results.xlsx"
        download_filepath = os.path.join(config.REPORTS_FOLDER, download_filename)

        with pd.ExcelWriter(download_filepath, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name='Analysis Results')
            
            worksheet = writer.sheets['Analysis Results']
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            alternating_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            # Style for wrapping text in the summary column
            wrap_alignment = Alignment(wrap_text=True, vertical='top')

            # Apply header style
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Apply alternating row color (this will be overridden by the review color if needed)
            for row_index in range(2, worksheet.max_row + 1):
                if row_index % 2 == 0:
                    for cell in worksheet[row_index]:
                        cell.fill = alternating_row_fill
            
            for idx, col in enumerate(final_df):
                column_letter = chr(65 + idx)
                series = final_df[col]
                max_len = max(
                    (series.astype(str).map(len).max(), len(str(series.name)))
                ) + 2
                
                # Set a max width for the Summary column, and a default for others
                if series.name == 'Summary' or series.name == 'Review Reason':
                    worksheet.column_dimensions[column_letter].width = min(max_len, 60)
                    # Apply text wrapping to all cells in the summary column
                    for cell in worksheet[column_letter]:
                        cell.alignment = wrap_alignment
                elif series.name == 'Needs Review':
                    worksheet.column_dimensions[column_letter].width = 15
                    # Apply a special color for 'Needs Review' cells
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet[f"{column_letter}{row}"]
                        if cell.value == 'Yes':
                            for cell in worksheet[row]:                                
                                cell.fill = PatternFill(start_color="FFF200", end_color="FFF200", fill_type="solid")
                else:
                    worksheet.column_dimensions[column_letter].width = min(max_len, 40)

        
        logger.info(f"Created styled downloadable report with review flags at: {download_filepath}")
        return download_filepath

    except Exception as e:
        logger.error(f"Error creating downloadable report for ID {report_id}: {e}", exc_info=True)
        return None
    

    history = utils.load_history()
    report_to_delete = None
    # Create a new list excluding the report to be deleted
    updated_history = [report for report in history if report['id'] != report_id]

    # Find the report entry to get file paths
    for report in history:
        if report['id'] == report_id:
            report_to_delete = report
            break
    
    if not report_to_delete:
        logger.warning(f"Attempted to delete non-existent report with ID: {report_id}")
        return False

    # If we found the report, proceed with deleting files
    try:
        # 1. Delete the main JSON report file
        if report_to_delete.get('file_path') and os.path.exists(report_to_delete['file_path']):
            os.remove(report_to_delete['file_path'])
            logger.info(f"Deleted JSON report: {report_to_delete['file_path']}")

        # 2. Delete the downloadable Excel report if it exists
        download_filename = f"{report_id}_analysis_results.xlsx"
        download_filepath = os.path.join(config.REPORTS_FOLDER, download_filename)
        if os.path.exists(download_filepath):
            os.remove(download_filepath)
            logger.info(f"Deleted downloadable report: {download_filepath}")

        # 3. Save the updated history file (with the entry removed)
        utils.save_history(updated_history)
        logger.info(f"Successfully deleted report entry with ID: {report_id}")
        return True

    except OSError as e:
        logger.error(f"Error deleting files for report {report_id}: {e}", exc_info=True)
        # Even if a file deletion fails, we still update the history to remove the broken entry
        utils.save_history(updated_history)
        return False